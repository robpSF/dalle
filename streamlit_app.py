import csv
import json
import openai
import openpyxl
import streamlit as sl
import pandas as pd
import re
import requests
from PIL import Image
import random
import math
from io import BytesIO
from geopy.geocoders import Nominatim
import datetime

geolocator = Nominatim(user_agent=sl.secrets["user_agent"])

# Choose which app to run
#selected_app = sl.selectbox("Choose the application to run", ["Personas", "PoL"])
#master_styles = ["Nikon D810", "generated in the style of photorealism, Sigma 85 mm f/1.4",
#                 "in the style of vector artwork", "in the style of flat art", "generated in a vintage style",
#                 "in the style of Banksy", "in the style of Roy Lichtenstein"]
#styles = sl.sidebar.multiselect("Pick acceptable image styles", master_styles)


# Here's all the functions -------------------------------------------------------
def list_purge(my_list):
    new_list = []
    for item in my_list:
        if len(item) > 3:
            point = item.find(".")
            new_list.append(item[point + 2:])
    return new_list


def get_subjects_and_bodies(lines):
    subjects = []
    bodies = []

    # sl.write(lines)
    nobody = True
    for line in lines:
        if line != "" and nobody:
            sub_position = line.find("Subject:")
            body_position = line.find("Body:")
            if line[sub_position:sub_position + 7] == "Subject":  # we're in business
                subjects.append(line[sub_position + 9:])
                nobody = True
            if line[body_position:body_position + 4] == "Body":  # we're in business
                nobody = False
                first_catch = True

        if line != "" and line.find("Subject:") > 0 and not nobody:
            bodies.append(body_text)
            nobody = True
            sub_position = line.find("Subject:")
            subjects.append(line[sub_position + 9:])

        if not nobody:
            # build the body!
            if first_catch:
                # start building
                body_position = line.find("Body:")
                body_text = line[body_position + 6:]
                first_catch = False
            else:
                sub_position = line.find("Subject:")
                if line == "":
                    line = "\r\n"
                    body_text = body_text + line
                elif line[sub_position:sub_position + 7] == "Subject":  # end of this body
                    bodies.append(body_text)
                    subjects.append(line[sub_position + 9:])
                    nobody = True
                else:
                    body_text = body_text + line
                # sl.write(body_text)

    # catch the last body!
    bodies.append(body_text)

    return subjects, bodies


def help_ticket(ticket_text):
    lines = ticket_text.split("\n")
    # first assume that Subjects and Bodies are on different lines
    subjects, bodies = get_subjects_and_bodies(lines)

    # if subjects is empty then "Body" is on the same line
    if subjects == []:
        subjects = []
        bodies = []
        for line in lines:
            if line != "" and line[0:7] == "Subject":
                body_position = line.find("Body:")
                subject = line[9:body_position]
                body = line[body_position + 2:]
                subjects.append(subject)
                bodies.append(body)

    sl.write(subjects)
    sl.write(bodies)
    return subjects, bodies


def get_emails(ticket_text):
    # do a quick scan to find "Body"

    subjects, bodies = help_ticket(ticket_text)
    sl.write(subjects)
    sl.write(bodies)
    return subjects, bodies


def get_persona_list(xl_sheet):
    personas = []
    for row in range(2, xl_sheet.max_row):
        personas.append(xl_sheet.cell(row, 1).value)

    return personas


def get_persona_list_from_df(df, faction_selected):
    # Filter the dataframe by the selected Faction
    filtered_df = df[df["Faction"] == faction_selected]

    personas = filtered_df["Name"].tolist()

    return personas


def convert_df(df):
    # IMPORTANT: Cache the conversion to prevent computation on every rerun
    return df.to_csv(index=False, line_terminator='\r\n').encode('utf-8')


# Load the OpenAI keys ------------------------------------------------------
# with open("config.json") as f:
#    config = json.load(f)

openai.organization = sl.secrets["organization"]
openai.api_key = sl.secrets["key"]
# ----------------------------------------------------------------------------
# URL for the DALL-E API
DALLE_API = "https://api.openai.com/v1/images/generations"
# DALL-E API Key
API_KEY = openai.api_key
openai.api_key = API_KEY

# Create the API request headers
headers = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {API_KEY}"
}
# this is a replacement image for when Dall-e throws an error
eggs = ["https://conducttr-a.akamaihd.net/fa/2302/05/22474/2023-02-05-10-00-02-100331.png",
        "https://conducttr-a.akamaihd.net/fa/2302/05/22474/2023-02-05-09-59-48-100330.png"]


# --- styles ---
# this array of style prompts helps create a better variety of Twitter bio pics

# --------------------------------------------

# Here's all the functions -------------------------------------------------------
def list_purge(my_list):
    new_list = []
    for item in my_list:
        if len(item) > 3:
            point = item.find(".")
            new_list.append(item[point + 2:])
    return new_list


def get_subjects_and_bodies(lines):
    subjects = []
    bodies = []

    # sl.write(lines)
    nobody = True
    for line in lines:
        if line != "" and nobody:
            sub_position = line.find("Subject:")
            body_position = line.find("Body:")
            if line[sub_position:sub_position + 7] == "Subject":  # we're in business
                subjects.append(line[sub_position + 9:])
                nobody = True
            if line[body_position:body_position + 4] == "Body":  # we're in business
                nobody = False
                first_catch = True

        if line != "" and line.find("Subject:") > 0 and not nobody:
            bodies.append(body_text)
            nobody = True
            sub_position = line.find("Subject:")
            subjects.append(line[sub_position + 9:])

        if not nobody:
            # build the body!
            if first_catch:
                # start building
                body_position = line.find("Body:")
                body_text = line[body_position + 6:]
                first_catch = False
            else:
                sub_position = line.find("Subject:")
                if line == "":
                    line = "\r\n"
                    body_text = body_text + line
                elif line[sub_position:sub_position + 7] == "Subject":  # end of this body
                    bodies.append(body_text)
                    subjects.append(line[sub_position + 9:])
                    nobody = True
                else:
                    body_text = body_text + line
                # sl.write(body_text)

    # catch the last body!
    bodies.append(body_text)

    return subjects, bodies


def help_ticket(ticket_text):
    lines = ticket_text.split("\n")
    # first assume that Subjects and Bodies are on different lines
    subjects, bodies = get_subjects_and_bodies(lines)

    # if subjects is empty then "Body" is on the same line
    if subjects == []:
        subjects = []
        bodies = []
        for line in lines:
            if line != "" and line[0:7] == "Subject":
                body_position = line.find("Body:")
                subject = line[9:body_position]
                body = line[body_position + 2:]
                subjects.append(subject)
                bodies.append(body)

    sl.write(subjects)
    sl.write(bodies)
    return subjects, bodies


def get_emails(ticket_text):
    # do a quick scan to find "Body"

    subjects, bodies = help_ticket(ticket_text)
    sl.write(subjects)
    sl.write(bodies)
    return subjects, bodies


def get_persona_list(xl_sheet):
    personas = []
    for row in range(2, xl_sheet.max_row):
        personas.append(xl_sheet.cell(row, 1).value)
    return personas


def get_persona_list_from_df(df, faction_selected):
    # Filter the dataframe by the selected Faction
    filtered_df = df[df["Faction"] == faction_selected]
    personas = filtered_df["Name"].tolist()
    return personas


def convert_df(df):
    # IMPORTANT: Cache the conversion to prevent computation on every rerun
    return df.to_csv(index=False, line_terminator='\r\n').encode('utf-8')


# Function to generate an image with DALL-E
def generate_dalle_image(prompt):
    # Build the API request payload
    data = """
    {
        """
    data += f'"model": "image-alpha-001",'
    data += f'"prompt": "{prompt}",'
    data += """
        "num_images":1,
        "size":"512x512",
        "response_format":"url"
    }
    """

    # Send the API request
    try:
        response = requests.post(DALLE_API, headers=headers, data=data)
        image_url = response.json()['data'][0]['url']
    except:
        sl.sidebar.write("Error with Dall-e")
        image_url = eggs[0]
    # Get the URL of the generated image

    # Get the image from the URL
    # image = Image.open(requests.get(image_url, stream=True).raw)

    return image_url

# Function to generate an image with DALL-E
def generate_dalle_image2(prompt):
    # Build the API request payload
    data = """
    {
        """
    data += f'"model": "image-alpha-001",'
    data += f'"prompt": "{prompt}",'
    data += """
        "num_images":1,
        "size":"512x512",
        "response_format":"url"
    }
    """

    # Send the API request
    try:
        response = requests.post(DALLE_API, headers=headers, data=data)
        sl.write(response)
        image_url = response.json()['data'][0]['url']
    except:
        sl.sidebar.write("Error with Dall-e")
        image_url = eggs[0]
    # Get the URL of the generated image

    # Get the image from the URL
    # image = Image.open(requests.get(image_url, stream=True).raw)

    return image_url


def generate_text(prompt, temp):
    response = openai.Completion.create(
        model="text-davinci-003",
        prompt=prompt,
        temperature=temp,
        max_tokens=1000,
        top_p=1.0,
        frequency_penalty=0.0,
        presence_penalty=0.0
    )
    # st.write(response)
    return response["choices"][0]["text"]


def strip_hashtags(text):
    return ' '.join(word for word in text.split() if not word.startswith("#"))


def bios_2_list(text):
    return [re.sub(r'[\[\]]', '', bio) for bio in re.findall(r'\[.*?\]', text)]


def repair_handle(s):
    parts = s.split("@")
    result = "".join(parts[1:])
    return result


def remove_emojis(input_string):
    emoji_pattern = re.compile("["
                               u"\U0001F600-\U0001F64F"  # emoticons
                               u"\U0001F300-\U0001F5FF"  # symbols & pictographs
                               u"\U0001F680-\U0001F6FF"  # transport & map symbols
                               u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
                               u"\U00002702-\U000027B0"
                               u"\U000024C2-\U0001F251"
                               "]+", flags=re.UNICODE)

    return emoji_pattern.sub(r'', input_string)


def generate_coordinates(latitude, longitude, radius, num_points):
    R = 6371  # Earth's radius in kilometers
    coordinates = []

    for i in range(num_points):
        theta = 2 * math.pi * random.random()
        r = random.uniform(0, radius)

        new_latitude = latitude + (r * math.cos(theta) / R)
        new_longitude = longitude + (r * math.sin(theta) / (R * math.cos(math.radians(latitude))))

        coordinates.append((new_latitude, new_longitude))

        # sl.write(str(new_latitude)+","+str(new_longitude))
    return coordinates


def set_the_map_coords(num_points):
    # coord_string = sl.sidebar.text_input("Enter coords e.g '51.5044672,-0.0821554'")
    # latitude, longitude = coord_string.split(",")

    coordinates = []
    coord_string = sl.sidebar.text_input("Paste Google maps URL here. This will be the center")
    if coord_string != "":
        latitude, longitude = coord_string.split("@")[1].split(",")[0:2]

        latitude = float(latitude)
        longitude = float(longitude)

        start_latitude = float(latitude)
        start_longitude = float(longitude)
        # start_latitude = 51.5044672
        # start_longitude = -0.0821554
        radius = sl.sidebar.slider("Radis", 10, 10000, 100)  # km

        coordinates = generate_coordinates(start_latitude, start_longitude, radius, num_points)
        # sl.write(coordinates)
        df = pd.DataFrame(coordinates, columns=['lat', 'lon'])
        sl.map(df)

        # sl.write(type(coordinates[0]))
    return coordinates


def fix_broken_names(names):
    if names.find(",") > 0:
        strip_hashtags(names)
        split_names = names.split(", ")
        # sl.sidebar.write(split_names)
        # Take the first name
        only_name = split_names[0].split(" ")[-1]
    else:
        only_name = names
    return only_name


def random_low_and_high(from_str, to_str):
    low = random.randint(int(from_str), int(to_str))
    high = random.randint(low, int(to_str))
    return low, high


def convert_df(df):
    # IMPORTANT: Cache the conversion to prevent computation on every rerun
    return df.to_csv(index=False, line_terminator='\r\n').encode('utf-8')


def get_town_and_country(location):
    # https://stackoverflow.com/questions/11390392/return-individual-address-components-city-state-etc-from-geopy-geocoder
    try:
        town = location.raw["address"]["city"]
    except:
        sl.sidebar.write("Address error, no city")
        sl.sidebar.write(location.raw)
        town = ""
    country = location.raw["address"]["country"]

    if town != "":
        if country != "":
            town_and_country = town + "," + country
        else:
            town_and_country = town
    else:
        if country != "":
            town_and_country = country
        else:
            town_and_country = " "

    return town_and_country


def random_date(start, end):
    return start + datetime.timedelta(
        seconds=random.randint(0, int((end - start).total_seconds())))


def to_excel(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    format1 = workbook.add_format({'num_format': '0.00'})
    worksheet.set_column('A:A', None, format1)
    writer.save()
    processed_data = output.getvalue()
    return processed_data


##########################################
#
#     POL GENERATOR
#
##########################################
def PoLGenerator(this_style):
    # Now read the persona file to get the list of personas -------------------------------------
    # Ask the user for the persona file
    sl.sidebar.write("___________________")
    sl.sidebar.write("CONFIGURATION PANEL")
    persona_file = sl.sidebar.file_uploader("Select persona file (.xlsx)")

    # If there's no persona file loaded yet then hang fire! --------------
    if persona_file != None:
        # Load the Excel file
        df = pd.read_excel(persona_file)
        # Get the unique values in the "Faction" column
        faction_options = df["Faction"].unique()
        # Create the select drop-down
        faction_selected = sl.sidebar.selectbox("Select a Faction", faction_options)

        # get the personas
        personas = get_persona_list_from_df(df, faction_selected)
        faction = faction_selected
        save_filename = re.sub(r'[^\w\s]', '_', faction)

        sl.sidebar.write("Enter the name of the PoL file you'll import later")
        filename = sl.sidebar.text_input(label="CSV filename", value=save_filename) + ".csv"

        sl.write("Using these personas...")
        sl.write(personas)

        # Ask about the channel ----------------
        # "Email" temporarily removed!
        channel = sl.sidebar.selectbox("Which channel?", ["Twitter", "Help Desk"])

        if channel == "Twitter":
            t_text = "Number of tweets"
            feed_text = "Enter Twitter feed prompt starting 'from {audience} about...'"
        elif channel == "Email":
            t_text = "Number of emails"
            feed_text = "Enter email feed prompt starting 'from {audience} about...'"
            to_persona = sl.sidebar.text_input(label="Enter is the email addressed to", value="who it may concern")
        elif channel == "Help Desk":
            t_text = "Number of support tickets"
            feed_text = "Enter support ticket feed prompt starting 'from {audience} about...'"

        temp = sl.sidebar.slider("Randomness of ideas", 0.0, 1.0, 0.9)
        if len(personas) == 1:
            num_tweets = 1
        else:
            num_tweets = sl.sidebar.slider(t_text, 1, len(personas), 2)

        scenario = sl.text_input(feed_text)
        image_prompt = sl.text_input("Enter image prompt", key=56) + " " + this_style[0]
        next_stage = sl.checkbox("generate PoL")
        do_images = sl.checkbox("generate and add images")

        # Now wait until everything has been set up and is ready to go -----------------------
        go = sl.button("do it!")
        if go:
            sl.sidebar.text("Drafting content")
            if do_images:
                image_url = generate_dalle_image(image_prompt)
                sl.image(image_url, width=200)

            if channel == "Twitter":
                my_prompt = "write " + str(num_tweets) + " Tweets " + scenario
            elif channel == "Email":
                my_prompt = "write " + str(
                    num_tweets) + " short emails including Subject and Body " + scenario + " to " + to_persona + " from 'THIS_PERSONA'"
            elif channel == "Help Desk":
                my_prompt = "write " + str(
                    num_tweets) + " abrupt support tickets including Subject and Body " + scenario

            response = openai.Completion.create(
                model="text-davinci-003",
                prompt=my_prompt,
                temperature=temp,
                max_tokens=1000,
                top_p=1.0,
                frequency_penalty=0.0,
                presence_penalty=0.0
            )

            # sl.write(response)
            AI_tweets = response["choices"][0]["text"]

            sl.subheader(channel)
            sl.write(AI_tweets)

            if next_stage:
                # create the dataframe that we'll export as CSV
                df = pd.DataFrame(
                    columns=['From', 'Faction', 'Method', 'Message', 'ImageURL', 'Subject', 'Subtitle', 'persona_img'])

                if channel == "Twitter":
                    # tidy up the OpenAI format
                    AI_tweets = AI_tweets.replace('"', '')
                    the_tweets = list_purge(AI_tweets.split('\n'))

                    sl.write(the_tweets)

                    # populate the dataframe
                    for i, item in enumerate(the_tweets):
                        if do_images:
                            image_url = generate_dalle_image(image_prompt)
                            sl.image(image_url, width=200)
                        else:
                            image_url = ""
                        new_data = {"From": personas[i], "Faction": faction, "Method": "Twitter", "Message": item,
                                    "ImageURL": image_url,
                                    "Subject": "", "Subtitle": "", "persona_img": ""}
                        df.loc[len(df)] = new_data

                if channel == "Help Desk":
                    subjects, bodies = help_ticket(AI_tweets)
                    # populate the dataframe
                    for i, item in enumerate(subjects):
                        new_data = {"From": personas[i], "Faction": faction, "Method": "Help Desk",
                                    "Message": bodies[i], "ImageURL": image_url,
                                    "Subject": subjects[i], "Subtitle": "", "persona_img": ""}
                        df.loc[len(df)] = new_data

                if channel == "Email":
                    subjects, bodies = get_emails(AI_tweets)
                    # populate the dataframe
                    for i, item in enumerate(subjects):
                        new_data = {"From": personas[i], "Faction": faction, "Method": "Help Desk",
                                    "Message": bodies[i], "ImageURL": image_url,
                                    "Subject": subjects[i], "Subtitle": "", "persona_img": ""}
                        df.loc[len(df)] = new_data

                # All done! Only left to download it!
                sl.write(df)
                csv = convert_df(df)

                sl.download_button(
                    label="Download data as CSV",
                    data=csv,
                    file_name=filename,
                    mime='text/csv',
                )
    return


# --------------END OF POL GENERATOR --------------

#####################################
#
#          PERSONA GENERATOR
#
#####################################
def persona_generator():
    sl.subheader("Persona Generator!")
    sl.sidebar.write("___________________")
    sl.sidebar.write("CONFIGURATION PANEL")

    temp = sl.sidebar.slider("Randomness of ideas", 0.0, 1.0, 0.9)
    num_peeps = sl.sidebar.slider("Number of personas", 1, 10, 2)

    coord_list = set_the_map_coords(num_peeps)
    do_images_too = sl.sidebar.checkbox("Generate images too", value=False)
    faction = sl.sidebar.text_input("Enter faction")
    dispostion = sl.sidebar.slider("Disposition", -5, 5, 1)
    tags = sl.sidebar.text_input("Enter tags")
    followers_from, followers_to = sl.sidebar.select_slider(
        'Select range of followers',
        options=['10', '100', '500', '1000', '4000', '5000', '10000', '30000', '100000', '1M'],
        value=('10', '100'))

    following_from, following_to = sl.sidebar.select_slider(
        'Select range of accounts followed',
        options=['10', '100', '500', '1000', '4000', '5000', '10000', '30000', '100000', '1M'],
        value=('100', '500'))

    tweets_from, tweets_to = sl.sidebar.select_slider(
        'Select range of tweets tweeted',
        options=['10', '100', '500', '1000', '4000', '5000', '10000'],
        value=('100', '500'))

    likes_from, likes_to = sl.sidebar.select_slider(
        'Select range of likes per tweet',
        options=['1', '5', '10', '50', '100', '200', '500', '1000', '5000'],
        value=('1', '5'))

    comments_from, comments_to = sl.sidebar.select_slider(
        'Select range of comments per tweet',
        options=['1', '5', '10', '50', '100', '200', '500', '1000', '5000'],
        value=('1', '5'))

    shares_from, shares_to = sl.sidebar.select_slider(
        'Select range of retweets per tweet',
        options=['1', '5', '10', '50', '100', '200', '500', '1000', '5000'],
        value=('1', '5'))

    sl.sidebar.write("Account date creation range")
    d_from = sl.sidebar.date_input("from date", datetime.date(2007, 6, 6))
    d_to = sl.sidebar.date_input("to date", datetime.date(2022, 12, 12))

    persona_brief = sl.text_input("Enter description of your personas")

    # Now wait until everything has been set up and is ready to go -----------------------
    go = sl.button("Generate bios")

    if go:
        twitter_bios = generate_text("[INSTRUCTIONS] write Twitter bios for " + str(
            num_peeps) + "based on the type of person described in {text}. The bios should start with [ and end with ].{text=}:" + persona_brief,
                                     temp)

        # sl.write(twitter_bios)
        bio_list = bios_2_list(twitter_bios)
        sl.write(bio_list)
        if do_images_too:
            # now to generate the images
            names = []
            handles = []
            images = []
            coord_strs = []

            text_col, image_col = sl.columns(2)

            i = 0
            sl.sidebar.text("Ok now building the profiles.")
            sl.sidebar.text(str(len(bio_list)) + " personas to create")

            for i, bio in enumerate(bio_list):
                sl.sidebar.write("Generating image " + str(i))
                bio_tweaked = remove_emojis(bio)

                name = generate_text("write one first and last name for " + bio_tweaked, temp)
                name = fix_broken_names(name)
                names.append(name)

                handle = repair_handle(generate_text("create a Twitter handle for " + bio_tweaked, temp))
                handle = strip_hashtags(handle)
                handle = remove_emojis(handle)
                handles.append(handle)

                this_style = styles[i % len(styles)]  # this uses modulo '%' to return to start when styles exhausted

                image_prompt = strip_hashtags(generate_text(
                    "[INSTRUCTIONS] summarise {text} into a good Dall-e prompt to make a profile picture." + tags + ". {text}=" + bio,
                    temp) + " " + this_style)
                # sl.write(image_prompt)
                image_url = generate_dalle_image(image_prompt)
                images.append(image_url)
                with text_col:
                    sl.subheader(name)
                    sl.write(handle)

                    account_created = random_date(d_from, d_to)
                    sl.write(account_created)

                    sl.write(bio)
                    this_coord = ", ".join("{:.7f}".format(num) for num in coord_list[i])
                    sl.write(this_coord)
                    coord_strs.append(this_coord)
                    location = geolocator.geocode(this_coord, addressdetails=True)
                    loc_str = get_town_and_country(location)
                    # sl.write(loc_str)
                    sl.write(this_style)
                with image_col:
                    sl.image(image_url, width=200)

                sl.write()

            # sl.write(names)
            # sl.write(handles)
            # sl.write(bio_list)
            # sl.write(coord_strs)
            # sl.write(images)

            # ok now ready to save as an Xls persona file
            columns = ["Name", "Handle", "Faction", "Disposition", "Tags", "RP", "Email", "Bio", "Goals", "Image",
                       "TwName",
                       "TwHandle", "TwVerified", "TwCreated", "TwBio", "TwProfileImg", "TwBgImg", "TwFollowers",
                       "TwFollowing", "TwPosts", "TwHistory", "TwWebsite", "TwCmtLo", "TwCmtHi", "TwLikeLo", "TwLikeHi",
                       "TwShareLo", "TwShareHi", "Location", "GPS", "URL"]
            # Create an empty dataframe with columns specified in the list
            df = pd.DataFrame(columns=columns)

            for i, bio in enumerate(bio_list):
                # get the followers, followed and tweet ranges
                TwFollowers_low, TwFollowers_high = random_low_and_high(followers_from, followers_to)
                TwFollowing_low, TwFollowing_high = random_low_and_high(following_from, following_to)
                Tweets_low, Tweets_high = random_low_and_high(tweets_from, tweets_to)
                likes_low, likes_high = random_low_and_high(likes_from, likes_to)
                comments_low, comments_high = random_low_and_high(comments_from, comments_to)
                shares_low, shares_high = random_low_and_high(shares_from, shares_to)

                df.loc[i] = [names[i], handles[i], faction, dispostion, tags, 0, "",
                             bio, "", images[i], names[i], handles[i], 0, account_created.strftime("%Y-%m-%d"),
                             bio, images[i], images[i], TwFollowers_low, TwFollowing_low, Tweets_low,
                             "",
                             "", comments_low, comments_high, likes_low, likes_high, shares_low, shares_high, loc_str,
                             coord_strs[i], ""]

            sl.write(df)
            # csv = convert_df(df)

            df_xlsx = to_excel(df)
            sl.download_button(label='📥 Download PERSONAS Result',
                               data=df_xlsx,
                               file_name=faction + '_personas_AI_generated.xlsx')
    return


# ---- end of persona generator ----------

######################################
#   HERE GOES!
######################################

password = sl.secrets["password"]
if password != sl.text_input("Enter password"): exit(600)

image_prompt = sl.text_input("Enter image prompt")
image_url = generate_dalle_image2(image_prompt)
sl.image(image_url)
